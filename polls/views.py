from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
import get_tokens

import os

os.chdir("C:\\Users\\Ranjith.Reddy\\PycharmProjects\\Django Projects\\data_analysis")
wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active
ws2=wb.create_sheet('count')
row_count = ws.max_row



def index(request):
        inc_count = 1
        f = open("list.txt", "a+")
        f.write("%d\r\n" % (inc_count))
        # Open the file back and read the contents
        contents = ws.cell(row=inc_count, column=1).value
        contents = contents.replace("_", " ")
        contents=contents.lower()
        tokens = get_tokens.get_tokens(contents)
        context = {'data_input': tokens , 'inc_count':inc_count, 'contents':contents}
        return render(request, 'polls/index.html', context)


def process_data(request,selected_tokens,inc_count):
    tokens_split=selected_tokens.split(",")
    count=0
    for i in range(1, row_count + 1):
        if inc_count == row_count:
            return render(request, 'polls/thankyou.html')
        content=ws.cell(row=i, column=1).value
        content=content.lower()
        if all (item in content for item in tokens_split):
           count=count+1
           ws.cell(row=i, column=1).value="none"
    row_next = ws2.max_row + 1
    ws2.cell(row=row_next, column=1).value=selected_tokens
    ws2.cell(row=row_next, column=2).value = count
    wb.save("test.xlsx")
    inc_count = inc_count + 1
    context = {'count':count , 'inc_count':inc_count }
    return render(request, 'polls/process_data.html', context)

def next_data(request,inc_count):
    if inc_count == row_count:
        return render(request, 'polls/thankyou.html')
    contents = ws.cell(row=inc_count, column=1).value
    contents=contents.lower()
    while (contents == "none"):

        if inc_count == row_count:
            return render(request, 'polls/thankyou.html')

        inc_count=inc_count+1
        contents= ws.cell(row=inc_count, column=1).value

    contents = contents.replace("_", " ")
    contents=contents.lower()
    tokens = get_tokens.get_tokens(contents)
    context = {'data_input': tokens, 'inc_count': inc_count, 'contents':contents}
    return render(request, 'polls/index.html', context)
